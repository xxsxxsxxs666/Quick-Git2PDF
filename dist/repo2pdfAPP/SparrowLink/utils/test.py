import torch
import matplotlib.pyplot as plt
from monai.inferers import sliding_window_inference
import os
import SimpleITK as sitk
from monai.data import Dataset, DataLoader, CacheDataset, decollate_batch
import numpy as np
from monai.transforms import (
    AsDiscrete,
    AsDiscreted,
    EnsureChannelFirstd,
    Compose,
    SaveImage,
)
from monai.metrics import HausdorffDistanceMetric
from monai.metrics import DiceMetric
from skimage.morphology import skeletonize, skeletonize_3d
import numpy as np
from openpyxl import Workbook, load_workbook
from monai.data import MetaTensor
import os
import SimpleITK as sitk
from utils.inferer import mirror_sliding_window_inference

def save_array(array, filename):
    array = array.detach().cpu().numpy()
    sitk.WriteImage(sitk.GetImageFromArray(array), filename)




def save_transformed_image(transformed_image, origin_image, space, save_path):
    image_after_transform = sitk.GetImageFromArray(np.array(transformed_image))
    image_after_transform.SetSpacing(space)
    image_after_transform.SetOrigin(origin_image.GetOrigin())
    image_after_transform.SetOrigin(origin_image.GetDirection())
    sitk.WriteImage(image_after_transform, save_path)


def cardio_vessel_segmentation_test(model, val_dataset, device, output_path, window_size, save_data=False):
    """
    Check best model output with the input image and label
    :param model: has loaded the stat_dic
    :param metric:
    :param val_loader:
    :param device:
    :return:
    """
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    val_loader = DataLoader(val_dataset, batch_size=1, num_workers=4)
    model.to(device)
    model.eval()
    save = Compose([
        AsDiscrete(argmax=True),
        SaveImage(output_dir=output_path, output_postfix='seg', print_log=False),
    ])
    save_origin = Compose([SaveImage(output_dir=output_path, output_postfix='label', print_log=False),])
    post_pred = Compose([AsDiscrete(argmax=True, to_onehot=2)])
    post_label = Compose([AsDiscrete(to_onehot=2)])
    metric_list = [DiceMetric(include_background=False, reduction="mean"),
                   HausdorffDistanceMetric(include_background=False, reduction="mean"), ]
    rows = [
        ('image_name', 'shape', str(metric_list[0].__class__).split(".")[-1][:-2], str(metric_list[1].__class__).split(".")[-1][:-2]),
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(rows[0])
    metric_path = os.path.join(output_path, 'metric.xlsx')
    wb.save(metric_path)
    with torch.no_grad():
        for i, val_data in enumerate(val_loader):
            roi_size = window_size
            sw_batch_size = 4
            cs_ms, val_labels = (
                val_data["image"].to(device),
                val_data["label"].to(device),
            )
            val_outputs = sliding_window_inference(
                cs_ms, roi_size, sw_batch_size, model
            )
            if save_data:
                for one_output, one_label in zip(decollate_batch(val_outputs), decollate_batch(val_labels)):
                    if not isinstance(one_output, MetaTensor):
                        meta = val_labels[0][0].meta_dict
                        one_output = MetaTensor(x=one_output, meta_dict=meta)
                    save(one_output)
                    save_origin(one_label)
            # use argmax
            arg_outputs = [post_pred(one_output) for one_output in decollate_batch(val_outputs)]
            val_labels = [post_label(one_label) for one_label in decollate_batch(val_labels)]
            # compute metric for current iteration
            row = [os.path.split(cs_ms[0][0].meta['filename_or_obj'])[-1], str(cs_ms[0][0].shape)]
            print(os.path.split(cs_ms[0][0].meta['filename_or_obj'])[-1]+':')
            for metric in metric_list:
                print_metric = metric(y_pred=arg_outputs, y=val_labels)
                print(f" {metric.__class__}:{print_metric.item()}", end='|')
                row.append(print_metric.item())
            print('')
            wb = load_workbook(metric_path)
            ws = wb.active
            ws.append(tuple(row))
            wb.save(metric_path)


            # # plot the slice
            # fig = plt.figure("check", (18, 6))
            # plt.subplot(1, 3, 1)
            # plt.title(f"image {i}")
            # plt.imshow(val_data["image"][0, 0, :, :, 80], cmap="gray")
            # plt.subplot(1, 3, 2)
            # plt.title(f"label {i}")
            # plt.imshow(val_data["label"][0, 0, :, :, 80])
            # plt.subplot(1, 3, 3)
            # plt.title(f"output {i}")
            # plt.imshow(torch.argmax(
            #     val_outputs, dim=1).detach().cpu()[0, :, :, 80])
            # plt.show()
            # image_name = val_data["image"][0][0].meta['filename_or_obj']
            # plt.savefig(os.path.join(output_path, (os.path.split(image_name)[-1]).split('.gz')[0]+'.jpg'))

        #  ------------save mean metric----------------- #
        row = ['mean', 'None']
        for metric in metric_list:
            print_metric = metric.aggregate()
            print(f" {metric.__class__}:{print_metric.item()}", end='|')
            row.append(print_metric.item())
        wb = load_workbook(metric_path)
        ws = wb.active
        ws.append(tuple(row))
        wb.save(metric_path)


def cardio_vessel_segmentation_multi_phase_test(model, val_dataset, device, output_path, window_size, save_data=False):
    """
    Check best model output with the input image and label
    :param model: has loaded the stat_dic
    :param metric:
    :param val_loader:
    :param device:
    :return:
    """
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    val_loader = DataLoader(val_dataset, batch_size=1, num_workers=4)
    model.to(device)
    model.eval()
    save = Compose([
        AsDiscrete(argmax=True),
        SaveImage(output_dir=output_path, output_postfix='seg', print_log=False),
    ])
    save_origin = Compose([SaveImage(output_dir=output_path, output_postfix='label', print_log=False),])
    post_pred = Compose([AsDiscrete(argmax=True, to_onehot=2)])
    post_label = Compose([AsDiscrete(to_onehot=2)])
    metric_list = [DiceMetric(include_background=False, reduction="mean"),
                   HausdorffDistanceMetric(include_background=False, reduction="mean"), ]
    rows = [
        ('image_name', 'shape', str(metric_list[0].__class__).split(".")[-1][:-2], str(metric_list[1].__class__).split(".")[-1][:-2]),
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(rows[0])
    metric_path = os.path.join(output_path, 'metric.xlsx')
    wb.save(metric_path)
    with torch.no_grad():
        for i, val_data in enumerate(val_loader):
            roi_size = window_size
            sw_batch_size = 4
            val_main, cs_a, val_labels = (
                val_data["main"].to(device),
                val_data["auxiliary"].to(device),
                val_data["label"].to(device),
            )
            val_outputs = sliding_window_inference(
                torch.cat([val_main, cs_a], dim=1), roi_size, sw_batch_size, model)
            if save_data:
                for one_output, one_label in zip(decollate_batch(val_outputs), decollate_batch(val_labels)):
                    if not isinstance(one_output, MetaTensor):
                        meta = val_labels[0][0].meta_dict
                        one_output = MetaTensor(x=one_output, meta_dict=meta)
                    save(one_output)
                    save_origin(one_label)
            # use argmax
            arg_outputs = [post_pred(one_output) for one_output in decollate_batch(val_outputs)]
            val_labels = [post_label(one_label) for one_label in decollate_batch(val_labels)]
            # compute metric for current iteration
            row = [os.path.split(val_main[0][0].meta['filename_or_obj'])[-1], str(val_main[0][0].shape)]
            print(os.path.split(val_main[0][0].meta['filename_or_obj'])[-1]+':')
            for metric in metric_list:
                print_metric = metric(y_pred=arg_outputs, y=val_labels)
                print(f" {metric.__class__}:{print_metric.item()}", end='|')
                row.append(print_metric.item())
            print('')
            wb = load_workbook(metric_path)
            ws = wb.active
            ws.append(tuple(row))
            wb.save(metric_path)


            # # plot the slice
            # fig = plt.figure("check", (18, 6))
            # plt.subplot(1, 3, 1)
            # plt.title(f"image {i}")
            # plt.imshow(val_data["image"][0, 0, :, :, 80], cmap="gray")
            # plt.subplot(1, 3, 2)
            # plt.title(f"label {i}")
            # plt.imshow(val_data["label"][0, 0, :, :, 80])
            # plt.subplot(1, 3, 3)
            # plt.title(f"output {i}")
            # plt.imshow(torch.argmax(
            #     val_outputs, dim=1).detach().cpu()[0, :, :, 80])
            # plt.show()
            # image_name = val_data["image"][0][0].meta['filename_or_obj']
            # plt.savefig(os.path.join(output_path, (os.path.split(image_name)[-1]).split('.gz')[0]+'.jpg'))

        #  ------------save mean metric----------------- #
        row = ['mean', 'None']
        for metric in metric_list:
            print_metric = metric.aggregate()
            print(f" {metric.__class__}:{print_metric.item()}", end='|')
            row.append(print_metric.item())
        wb = load_workbook(metric_path)
        ws = wb.active
        ws.append(tuple(row))
        wb.save(metric_path)


def save_metadata(metadata, output_path, output_postfix=''):
    """
    save the metadata of the dataset
    :return:
    """
    save = Compose([SaveImage(output_dir=output_path, output_postfix=output_postfix, print_log=False, ), ])
    save(metadata)


def cardio_vessel_segmentation_multi_phase_with_image_test(model, key, val_dataset, device, output_path, window_size, save_data=False):
    """
    Check best model output with the input image and label
    :param model: has loaded the stat_dic
    :param key: choose the information for input
    :param metric:
    :param val_loader:
    :param device:
    :return:
    """
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    val_loader = DataLoader(val_dataset, batch_size=1, num_workers=4)
    model.to(device)
    model.eval()
    metric_list = [DiceMetric(include_background=True, reduction="mean"),
                   HausdorffDistanceMetric(include_background=True, reduction="mean"), ]
    rows = [
        ('image_name', 'shape',
         'gt_' + str(metric_list[0].__class__).split(".")[-1][:-2],
         str(metric_list[0].__class__).split(".")[-1][:-2],
         'gt_' + str(metric_list[1].__class__).split(".")[-1][:-2],
         str(metric_list[1].__class__).split(".")[-1][:-2],),
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(rows[0])
    metric_path = os.path.join(output_path, 'metric.xlsx')
    wb.save(metric_path)
    with torch.no_grad():
        for i, val_data in enumerate(val_loader):
            cs_m = val_data['CS_M'].to(device)
            cs_dl = val_data['CS_DL'].to(device)
            cs_dlgt = val_data['CS_DLGT'].to(device)
            cs_a = val_data['CS_A'].to(device)
            val_labels = val_data['label'].to(device)
            cs_ms = torch.zeros((val_data[key[0]].shape[0], len(key), val_data[key[0]].shape[2],
                                      val_data[key[0]].shape[3], val_data[key[0]].shape[4]), device=device)
            for j in range(len(key)):
                cs_ms[:, j, :, :, :] = val_data[key[j]]
            roi_size = window_size
            sw_batch_size = 4
            val_outputs = sliding_window_inference(cs_ms, roi_size, sw_batch_size, model)
            val_outputs = torch.argmax(val_outputs, dim=1, keepdim=True)
            rcs_gt = val_outputs * cs_dlgt + cs_m * (1 - cs_dlgt)
            rcs_gt = [one_output for one_output in decollate_batch(rcs_gt)]
            rcs = val_outputs * cs_dl + cs_m * (1 - cs_dl)
            rcs = [one_output for one_output in decollate_batch(rcs)]
            val_outputs = [one_output for one_output in decollate_batch(val_outputs)]
            val_labels = [one_label for one_label in decollate_batch(val_labels)]
            cs_m = [one_input for one_input in decollate_batch(cs_m)]
            cs_a = [one_input for one_input in decollate_batch(cs_a)]
            cs_dl = [one_input for one_input in decollate_batch(cs_dl)]
            cs_dlgt = [one_input for one_input in decollate_batch(cs_dlgt)]

            if save_data:
                for i, one_output in enumerate(val_outputs):
                    if not isinstance(one_output, MetaTensor):
                        meta = val_labels[i].meta
                        one_output = MetaTensor(x=one_output, meta=meta)
                    save_metadata(one_output, output_path, output_postfix="RS")
                for one_output in rcs_gt:
                    save_metadata(one_output, output_path, output_postfix="RCSGT")
                for one_output in rcs:
                    save_metadata(one_output, output_path, output_postfix="RCS")
                for one_label in val_labels:
                    save_metadata(one_label, output_path, output_postfix="GT")
                for one_broken in cs_dl:
                    save_metadata(one_broken, output_path, output_postfix="CS_DL")
                for one_broken in cs_dlgt:
                    save_metadata(one_broken, output_path, output_postfix="CS_DLGT")
                for one_input in cs_m:
                    save_metadata(one_input, output_path, output_postfix="CS_M")
                for one_input in cs_a:
                    save_metadata(one_input, output_path, output_postfix="CS_A")

            # compute metric for current iteration
            row = [os.path.split(val_data['CS_M'][0][0].meta['filename_or_obj'])[-1], str(val_data['CS_M'][0][0].shape)]
            print(os.path.split(val_data['CS_M'][0][0].meta['filename_or_obj'])[-1]+':')
            for metric in metric_list:
                print_metric_gt = metric(y_pred=rcs_gt, y=val_labels)
                print_metric = metric(y_pred=rcs, y=val_labels)
                print(f" {metric.__class__}:{print_metric_gt.item()}, {print_metric.item()}", end='|')
                row.append(print_metric_gt.item())
                row.append(print_metric.item())
            print('')
            wb = load_workbook(metric_path)
            ws = wb.active
            ws.append(tuple(row))
            wb.save(metric_path)


            # # plot the slice
            # fig = plt.figure("check", (18, 6))
            # plt.subplot(1, 3, 1)
            # plt.title(f"image {i}")
            # plt.imshow(val_data["image"][0, 0, :, :, 80], cmap="gray")
            # plt.subplot(1, 3, 2)
            # plt.title(f"label {i}")
            # plt.imshow(val_data["label"][0, 0, :, :, 80])
            # plt.subplot(1, 3, 3)
            # plt.title(f"output {i}")
            # plt.imshow(torch.argmax(
            #     val_outputs, dim=1).detach().cpu()[0, :, :, 80])
            # plt.show()
            # image_name = val_data["image"][0][0].meta['filename_or_obj']
            # plt.savefig(os.path.join(output_path, (os.path.split(image_name)[-1]).split('.gz')[0]+'.jpg'))

        #  ------------save mean metric----------------- #
        row = ['mean', 'None']
        for metric in metric_list:
            print_metric = metric.aggregate()
            print(f" {metric.__class__}:{print_metric.item()}", end='|')
            row.append(print_metric.item())
        wb = load_workbook(metric_path)
        ws = wb.active
        ws.append(tuple(row))
        wb.save(metric_path)